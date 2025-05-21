# 사전 설치 : pip install flask pymysql
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from db import Database
from datetime import datetime ,date
import pandas as pd 
import io 
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from openpyxl.styles import Alignment
import atexit   # 애플리케이션 종료시 실행을 요청 (ex. DB연결 종료)

app = Flask(__name__)   # Flask 앱 초기화
db = Database()   # DB 초기화
conn = db.connection
cursor = conn.cursor()

@app.route('/main')
def main():
    return render_template('main.html')  # HTML 파일 이름
@app.route('/registration')
def index():
    return render_template('registration.html')  # HTML 파일 이름
@app.route('/user')
def users_page():
    return render_template('user.html')
@app.route('/rentals')
def test_page():
    return render_template('rentals.html')
#-------------------  main ---------------------
@app.route('/status-data')
def status_data():
    today = date.today()
    # db.connection과 cursor를 사용한다고 가정
    conn = db.connection
    cursor = conn.cursor()
    
    # 오늘 대여 수
    cursor.execute("SELECT COUNT(*) AS cnt FROM rentals WHERE DATE(rent_date) = %s", (today,))
    today_rentals = cursor.fetchone()['cnt']

    # 오늘 반납 수
    cursor.execute("SELECT COUNT(*) AS cnt FROM rentals WHERE return_date = %s", (today,))
    today_returns = cursor.fetchone()['cnt']

    # 누적 대여 금액
    cursor.execute("SELECT SUM(price) AS total FROM rentals WHERE price IS NOT NULL")
    total_revenue = cursor.fetchone()['total'] or 0

    # 인기 대여 품목 Top 3
    cursor.execute("""
        SELECT i.name
        FROM rentals r
        JOIN items i ON r.item_id = i.item_id
        GROUP BY i.name
        ORDER BY COUNT(*) DESC
        LIMIT 3
    """)
    top_items = [row['name'] for row in cursor.fetchall()]

    # 오늘 대여 수가 0이면 메시지로 변환
    if today_rentals == 0:
        today_rentals = "오늘 대여 수 없습니다"

    return jsonify({
        "today_rentals": today_rentals,
        "today_returns": today_returns,
        "total_revenue": total_revenue,
        "top_items": top_items
    })
# ------------------ 사용자 관련 ------------------
@app.route('/users', methods=['POST'])
def create_user():
    name = request.json.get('name')
    if not name:
        return jsonify({'error': '이름을 입력하세요'}), 400
    cursor.execute("INSERT INTO users (name) VALUES (%s)", (name,))
    conn.commit()
    return jsonify({'message': '사용자 등록 완료'}), 201

@app.route('/users', methods=['GET'])
def get_users():
    cursor.execute("SELECT * FROM users")
    users = cursor.fetchall()
    return jsonify(users)

@app.route('/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    new_name = request.json.get('name')
    cursor.execute("UPDATE users SET name = %s WHERE user_id = %s", (new_name, user_id))
    conn.commit()
    return jsonify({'message': '사용자 정보 수정 완료'})

@app.route('/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    cursor.execute("DELETE FROM users WHERE user_id = %s", (user_id,))
    conn.commit()
    return jsonify({'message': '사용자 삭제 완료'})

# ------------------ 물품 관련 ------------------

@app.route('/items', methods=['POST'])
def create_item():
    data = request.json
    name = data.get('name')
    price = data.get('price')
    quantity = data.get('quantity', 1)

    if not name or price is None:
        return jsonify({'error': '이름과 가격은 필수입니다'}), 400

    cursor.execute(
        "INSERT INTO items (name, price, quantity) VALUES (%s, %s, %s)",
        (name, price, quantity)
    )
    conn.commit()
    return jsonify({'message': '물품 등록 완료'}), 201

@app.route('/items', methods=['GET'])
def get_items():
    cursor.execute("SELECT * FROM items")
    items = cursor.fetchall()
    return jsonify(items)

@app.route('/items/<int:item_id>', methods=['PUT'])
def update_item(item_id):
    data = request.json
    name = data.get('name')
    price = data.get('price')
    quantity = data.get('quantity')

    cursor.execute(
        "UPDATE items SET name = %s, price = %s, quantity = %s WHERE item_id = %s",
        (name, price, quantity, item_id)
    )
    conn.commit()
    return jsonify({'message': '물품 수정 완료'})

@app.route('/items/<int:item_id>', methods=['DELETE'])
def delete_item(item_id):
    cursor.execute("DELETE FROM items WHERE item_id = %s", (item_id,))
    conn.commit()
    return jsonify({'message': '물품 삭제 완료'})

# 대여 현황 조회
@app.route('/api/rentals')
def get_rentals():
    query = """
        SELECT r.rental_id, r.user_id, r.item_id, r.rent_date, r.price, r.status,
        u.name AS user_name, i.name AS item_name
        FROM rentals r
        JOIN users u ON r.user_id = u.user_id
        JOIN items i ON r.item_id = i.item_id
    """
    cursor.execute(query)
    rentals = cursor.fetchall()
    # rentals 리스트를 JSON으로 반환
    return jsonify(rentals)

# 대여 등록
@app.route('/api/rentals', methods=['POST'])
def create_rental():
    data = request.json
    user_id = data.get('user_id')
    item_id = data.get('item_id')
    rent_date = data.get('rent_date')
    price = data.get('price')
    status = data.get('status', '대여중')

    cursor.execute("""
        INSERT INTO rentals (user_id, item_id, rent_date, price, status)
        VALUES (%s, %s, %s, %s, %s)
    """, (user_id, item_id, rent_date, price, status))
    conn.commit()
    return jsonify({'message': '대여 등록 완료'}), 201

# 반납 처리
@app.route('/api/rentals/<int:rental_id>/return', methods=['PUT'])
def return_rental(rental_id):
    try:
        return_date = datetime.now()
        cursor.execute("""
            UPDATE rentals
            SET status = '반납완료', return_date = %s
            WHERE rental_id = %s
        """, (return_date, rental_id))
        conn.commit()
        return jsonify({'message': '반납 처리 완료'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
# __________________________________________실적현황____________________________________________
@app.route('/status', methods=['GET'])
def status():
    data = Database.fetch_rental_performance()
    
    if not data:
        return render_template('status.html',
                             total_rentals=0,
                             total_revenue=0,
                             user_stats=[],
                             item_stats=[])

    return render_template('status.html',
                         total_rentals=data['total_rentals'],
                         total_revenue=data['total_revenue'],
                         user_stats=data['user_stats'],
                         item_stats=data['item_stats'])

@app.route('/api/monthly-stats')
def monthly_stats():
    data = Database.fetch_monthly_stats()
    
    if not data:
        return jsonify({'success': False})
        
    return jsonify({'success': True, 'data': data})
# __________________________________________반납 아이템 목록____________________________________________
#반납 아이템 목록
@app.route('/return_item', methods=['GET'])
def get_return_item():
    try:
        with db.connection.cursor() as cursor:
            # rentals.status 에서 '반납 완료'인 대여만, 그에 해당하는 items 정보를 가져옵니다
            cursor.execute("""
                SELECT 
                    i.item_id, i.name, i.price, i.quantity, i.created_at,
                    r.status, r.return_date, r.item_condition AS item_condition
                FROM rentals r
                JOIN items i ON r.item_id = i.item_id
                WHERE r.status = '반납완료'
            """)
            items = cursor.fetchall()
            print(items)
        return render_template('return_item.html', items=items)
        
    except Exception as e:
        return str(e), 500

#반납 목록 엑셀 다운로드 (상단 import 확인 필)
#pip install pandas openpyxl 필요

@app.route('/download_returns')
def download_returns():
    try:
        query = """
                SELECT 
                    i.item_id, i.name, i.price, i.quantity, i.created_at,
                    r.status, r.return_date, r.item_condition AS item_condition
                FROM rentals r
                JOIN items i ON r.item_id = i.item_id
                WHERE r.status = '반납완료'
            """
        
        with db.connection.cursor() as cursor:
            cursor.execute(query)
            results = cursor.fetchall()
            
            if not results:
                return "반납완료 상태의 데이터가 없습니다", 204

        # 컬럼명 정의
        df = pd.DataFrame(results)
        df.rename(columns={
            'item_id':'아이템ID', 
            'name':'이름', 
            'price':'가격', 
            'quantity':'수량', 
            'created_at':'등록일', 
            'status':'대여 상태', 
            'return_date':'반납일시', 
            'item_condition':'물품상태'
        }, inplace=True)

        # 메모리 버퍼에 저장
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='반납리스트')
            
            worksheet = writer.sheets['반납리스트']
            
            column_widths = {
                '아이템ID': 10,
                '이름': 20,
                '가격': 12,
                '수량': 8,
                '등록일': 25,
                '대여 상태': 12,
                '반납일시': 25,
                '물품상태': 12
            }
            
            for col_num, col_name in enumerate(df.columns, 1):  # 1-based index
                col_letter = get_column_letter(col_num)
                width = column_widths.get(col_name, 15)
                worksheet.column_dimensions[col_letter].width = width
                
                # 날짜 형식 지정 (등록일과 반납일시 컬럼)
            date_fmt = numbers.FORMAT_DATE_DATETIME  # 'yyyy-mm-dd hh:mm:ss'
            for row in range(2, len(df) + 2):  # 데이터 행들 (엑셀은 1부터, 1행은 헤더)
                worksheet[f'E{row}'].number_format = date_fmt  # 등록일 열(E)
                worksheet[f'G{row}'].number_format = date_fmt  # 반납일시 열(G)
                
            alignment = Alignment(horizontal='center', vertical='center')

            max_row = len(df) + 1  # 데이터 행 수 + 헤더 1행
            max_col = len(df.columns)

            for row in range(1, max_row + 1):  # 헤더 포함
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = alignment
       
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name='반납리스트.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
# ------------------ 종료 시 DB 연결 닫기 ------------------
import atexit
atexit.register(db.close)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)