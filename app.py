from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from datetime import datetime
import pandas as pd       # DataFrame 및 엑셀 저장용
import io                 # 메모리 버퍼에 엑셀을 저장하기 위함
from db import Database   # 사용자 정의 DB 연결 클래스
import pymysql
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from openpyxl.styles import Alignment

app = Flask(__name__)
db = Database()

#반납 아이템 목록
@app.route('/return_item', methods=['GET'])
def get_return_item():
    try:
        with db.connection.cursor() as cursor:
            # rentals.status 에서 '반납 완료'인 대여만, 그에 해당하는 items 정보를 가져옵니다
            cursor.execute("""
                SELECT 
                    i.item_id, i.name, i.price, i.quantity, i.created_at,
                    r.status, r.return_date, r.item_condition AS item_condition
                FROM rentals r
                JOIN items i ON r.item_id = i.item_id
                WHERE r.status = '반납완료'
            """)
            items = cursor.fetchall()
            print(items)
        return render_template('return_item.html', items=items)
        
    except Exception as e:
        return str(e), 500

#반납 목록 엑셀 다운로드 (상단 import 확인 필)
#pip install pandas openpyxl 필요

@app.route('/download_returns')
def download_returns():
    try:
        query = """
                SELECT 
                    i.item_id, i.name, i.price, i.quantity, i.created_at,
                    r.status, r.return_date, r.item_condition AS item_condition
                FROM rentals r
                JOIN items i ON r.item_id = i.item_id
                WHERE r.status = '반납완료'
            """
        
        with db.connection.cursor() as cursor:
            cursor.execute(query)
            results = cursor.fetchall()
            
            if not results:
                return "반납완료 상태의 데이터가 없습니다", 204

        # 컬럼명 정의
        df = pd.DataFrame(results)
        df.rename(columns={
            'item_id':'아이템ID', 
            'name':'이름', 
            'price':'가격', 
            'quantity':'수량', 
            'created_at':'등록일', 
            'status':'대여 상태', 
            'return_date':'반납일시', 
            'item_condition':'물품상태'
        }, inplace=True)

        # 메모리 버퍼에 저장
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='반납리스트')
            
            worksheet = writer.sheets['반납리스트']
            
            column_widths = {
                '아이템ID': 10,
                '이름': 20,
                '가격': 12,
                '수량': 8,
                '등록일': 25,
                '대여 상태': 12,
                '반납일시': 25,
                '물품상태': 12
            }
            
            for col_num, col_name in enumerate(df.columns, 1):  # 1-based index
                col_letter = get_column_letter(col_num)
                width = column_widths.get(col_name, 15)
                worksheet.column_dimensions[col_letter].width = width
                
                # 날짜 형식 지정 (등록일과 반납일시 컬럼)
            date_fmt = numbers.FORMAT_DATE_DATETIME  # 'yyyy-mm-dd hh:mm:ss'
            for row in range(2, len(df) + 2):  # 데이터 행들 (엑셀은 1부터, 1행은 헤더)
                worksheet[f'E{row}'].number_format = date_fmt  # 등록일 열(E)
                worksheet[f'G{row}'].number_format = date_fmt  # 반납일시 열(G)
                
            alignment = Alignment(horizontal='center', vertical='center')

            max_row = len(df) + 1  # 데이터 행 수 + 헤더 1행
            max_col = len(df.columns)

            for row in range(1, max_row + 1):  # 헤더 포함
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = alignment
       
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name='반납리스트.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    

if __name__ == '__main__':
    app.run(debug=True)
    
    #최신
